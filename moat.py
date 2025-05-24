from dotenv import load_dotenv
import os
import datetime as dt
from openpyxl import load_workbook
import textwrap
from queue import Queue
import threading
import time
import polars as pl
from google import genai
from google.genai.types import Tool, GenerateContentConfig, GoogleSearch
import shelve

CACHE_FILE = "company_cache"

model_id = "gemini-2.5-flash-preview-04-17" # 05-20

google_search_tool = Tool(
    google_search = GoogleSearch()
)

# Use multithreading to speed up
num_threads = 5 #5 worked just fine for limit=50

print(model_id + ' AI model working in progress.. May take up to few minutes.')

# today = dt.datetime.today().weekday()

data = []
data_lock = threading.Lock()

moat = {
    3: "Unbreachable(+3)",
    2: "Strong(+2)",
    1: "Narrow(+1)",
    0: "None(+0)"
}

# Load environment variables from .env file
load_dotenv()

# Get the API key
api_key = os.getenv("GEMINI_API_KEY")

# Use it with OpenAI
client = genai.Client(api_key=api_key)

q = Queue()
with shelve.open(CACHE_FILE) as cache:
    for name, score in cache.items():
        q.put((name, score))


def process_moat():
    while not q.empty():
        (name, score) = q.get()
        try:
            # - Return a single integer as the response output without any text explanation

            user_prompt = f"""

            You are a financial analysis AI trained in the style of Warren Buffett's long-term investment philosophy. 
            Your mission is to use **web search capabilities** to analyze the long-term competitive advantage (economic moat) of {name}, 
            strictly and ONLY based on **credible, reputable, and fact-based news sources published after 2024**. 
            Do not use your own knowledge for the analysis. The analysis must be evidence-based

            **Before conducting the analysis, you must use the web tool to search for the most recent data (must be published after 2024) ONLY from the following designated sources:**
            * Bloomberg, Reuters, The Wall Street Journal, CNBC, Yahoo Finance, Korea Economic Daily, Maeil Business Newspaper, YTN, Yonhap News, Chosun Ilbo, JoongAng Ilbo, 
            Financial Times, The Economist, Nikkei Asia, OR any other highly credidble finance/business news sources

            **Please cite the source and its published date in parentheses.**
            Example: (Bloomberg, 2025.05.22)

            **Refrain from collecting all the information from a single source. Try your best to use multiple sources.**

            **Do not use any data that does not clearly cite a source.**

            **You must entirely respond in Korean.

            ---

            ### Criteria for Analysis:

            1. Brand strength
            2. Network effects
            3. Cost advantage
            4. Switching costs
            5. Intangible assets (e.g., patents, licenses)
            6. Economies of scale and market dominance

            ---

            ### Investigate the following factors:

            * **Key assets** (e.g., patents, ecosystem, strong brand)
            * **Customer base** (e.g., mass market, enterprises, governments)
            * **Consistency and durability** (e.g., can this advantage last for decades? Is it resilient across economic cycles?)
            * **Cash flow and free cash flow**
            * **Customer loyalty and pricing power** (e.g., does the customer prefer the product even at higher prices?)
            * **Quality of management** (Buffett favors “able and honest” managers who align with shareholders and allocate capital wisely)

            ---

            ### Evaluation Criteria:

            1. Type of moat
            2. Durability of the moat
            3. Key risks or threats
            4. Final rating:

            * `3`: **Unbreachable** : Extremely rare; based on monopolistic position, proprietary technology, or powerful network effects
            * `2`: **Strong** : Sustainable and hard-to-replicate competitive advantage
            * `1`: **Narrow** : Some edge, but likely to weaken over time
            * `0`: **No moat** : Little or no sustainable advantage; easily exposed to competition

            """.strip()

            response = client.models.generate_content(
            model=model_id,
            config=GenerateContentConfig(
                tools=[google_search_tool],
                response_modalities=["TEXT"],
            ),
            contents=user_prompt)

            # 줄 단위로 나눈 후, 각각에 대해 줄바꿈 적용
            wrapped_lines = []
            for line in response.text.strip().split('\n'):
                # 줄의 시작 공백(들여쓰기) 보존
                leading_spaces = len(line) - len(line.lstrip(' '))
                indent = ' ' * leading_spaces

                # bullet 유지되도록 첫 단어 확인
                if line.lstrip().startswith(("-", "*", "•")) or line.lstrip()[:2].isdigit():
                    first_word = line.split()[0]
                    rest = ' '.join(line.split()[1:])
                    wrapped = textwrap.fill(rest, width=120 - len(indent) - len(first_word) - 1,
                                            initial_indent=indent + first_word + ' ',
                                            subsequent_indent=indent + ' ' * (len(first_word) + 1))
                else:
                    wrapped = textwrap.fill(line, width=120,
                                            initial_indent=indent,
                                            subsequent_indent=indent)
                wrapped_lines.append(wrapped)

            # 최종 텍스트
            ai_generated_moat = '\n'.join(wrapped_lines)

            result = {
                "기업": name,
                "점수(9)": score,
                "분석": ai_generated_moat,
            }

            with data_lock:
                data.append(result)

        except Exception as e:
            if "429" in str(e):
                print("Too many requests! Waiting 10 seconds...")
                time.sleep(10)
            data.append({
                "기업": name,
                "점수(9)": 0,
                "분석": '',
            })

        finally:
            q.task_done()
            time.sleep(2)
    

threads = []

for _ in range(num_threads):
    t = threading.Thread(target=process_moat)
    t.start()
    threads.append(t)

for t in threads:
    t.join()

df = pl.DataFrame(data)

df_sorted = df.sort("점수(9)", descending = True)

df_sorted.to_pandas().to_excel("moat_analysis.xlsx", index=False)

wb = load_workbook("moat_analysis.xlsx")
ws = wb.active

for col in ['C']:
    ws.column_dimensions[col].width = 200

for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 500

wb.save("moat_analysis.xlsx")